"""XLSX 셀을 마크다운 표로 뽑는다.

RAG Engine 기본 파서는 XLSX 를 읽지 못한다. 원본을 그대로 올려봐야 검색되지
않으므로 지금까지는 경로 사이드카(`.meta.md`) 한 줄만 색인했는데, 그러면
"그 표 안에 뭐가 있냐"는 질문에 구조적으로 답이 안 나온다. 실제로 운영 중
그 질문이 들어와 9회 재질의 끝에 실패한 사례가 있다(2026-07-29 09:21).

셀 좌표와 값은 파일에 그대로 들어 있어 추론할 것이 없다. HWP 처럼 레이아웃에서
표를 알아내는 작업이 아니라, 읽어서 옮겨 적는 작업이다.

측정(운영 코퍼스 xlsx 116건 전량):
    89건  진짜 OOXML    → 87건 즉시 성공, 2건은 styles.xml 결함(아래 복구 경로)
    27건  암호 걸린 파일 → 열 수 없음. XlsxParseError 로 알리고 호출측이
                          사이드카만 남기던 기존 동작으로 떨어진다
"""

from __future__ import annotations

import datetime as _dt
import io
import logging
import re
import zipfile
from xml.etree import ElementTree as ET

logger = logging.getLogger(__name__)

# 병적으로 큰 시트에서 메모리/시간이 터지지 않게 두는 상한.
# 운영 코퍼스 최대가 146,472 셀이라 실측 최대의 2배로 잡았다.
MAX_CELLS = 300_000

# 출력 바이트 상한. 셀 수만 제한하면 부족하다 — 셀 30만개가 29MB 를 만들어
# RAG 한도(10MB)를 넘긴 사례가 있었다(2026-07-29 재색인). 한도를 넘기면 색인
# 자체가 실패해 사이드카로도 못 찾게 되므로, 잘라서라도 넣는 편이 낫다.
# 머리말과 한도 사이 여유를 두고 8MB.
MAX_BYTES = 8 * 1024 * 1024

# 암호화된 OOXML 은 OLE2 컨테이너로 감싸여 구형 .xls 와 매직바이트가 같다.
# (MS-OFFCRYPTO) 확장자로는 구분되지 않아 바이트로 본다.
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0"

# `<x:cellStyles>` 처럼 네임스페이스 접두사를 붙여 쓰는 파일이 있어 접두사를 함께 받는다.
# (접두사 없는 형태만 찾으면 정작 문제 있는 파일을 못 잡는다)
_CELL_STYLES_RE = re.compile(
    rb"<(?:\w+:)?cellStyles\b.*?</(?:\w+:)?cellStyles>|<(?:\w+:)?cellStyles\b[^>]*/>",
    re.S,
)


class XlsxParseError(RuntimeError):
    """열 수 없는 XLSX — 암호, 손상 등."""


def _fmt(value: object) -> str:
    """셀 값을 표에 넣을 문자열로. 숫자 1.0 이 '1.0' 으로 새는 것을 막는다."""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.isoformat(sep=" ") if isinstance(value, _dt.datetime) else value.isoformat()
    text = str(value)
    # 표가 깨지지 않도록. 셀 안 줄바꿈은 공백으로 편다.
    return text.replace("|", r"\|").replace("\r", " ").replace("\n", " ").strip()


def _strip_named_styles(data: bytes) -> bytes:
    """`xl/styles.xml` 의 cellStyles 를 들어낸 사본을 만든다.

    `<cellStyle xfId="0" builtinId="0"/>` 처럼 name 속성이 빠진 항목이 있으면
    openpyxl 이 로드 중 TypeError 로 죽는다. 셀 값과는 무관한 스타일 메타라,
    통째로 지우고 다시 읽으면 값은 온전하다.
    """
    src = io.BytesIO(data)
    out = io.BytesIO()
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            payload = zin.read(item.filename)
            if item.filename == "xl/styles.xml":
                payload = _CELL_STYLES_RE.sub(b"", payload)
            zout.writestr(item, payload)
    return out.getvalue()


def _load(data: bytes):
    import openpyxl

    try:
        return openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except TypeError as exc:
        # styles.xml 결함 — 스타일을 들어내고 한 번 더.
        logger.info("xlsx styles 복구 시도: %s", exc)
        try:
            repaired = _strip_named_styles(data)
        except Exception as inner:  # noqa: BLE001
            raise XlsxParseError(f"styles 복구 실패: {inner}") from exc
        try:
            return openpyxl.load_workbook(
                io.BytesIO(repaired), read_only=True, data_only=True
            )
        except Exception as inner:  # noqa: BLE001
            raise XlsxParseError(f"복구 후에도 열리지 않음: {inner}") from exc
    except zipfile.BadZipFile as exc:
        raise XlsxParseError(f"ZIP 아님(손상 추정): {exc}") from exc
    except Exception as exc:  # noqa: BLE001
        raise XlsxParseError(f"{type(exc).__name__}: {exc}") from exc


_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_REL = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
_PKG_REL = "{http://schemas.openxmlformats.org/package/2006/relationships}"
_CELL_REF = re.compile(r"^([A-Z]+)(\d+)$")


def _col_num(letters: str) -> int:
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch) - 64)
    return n


def _sheet_xml_paths(zf: zipfile.ZipFile) -> list[str]:
    """workbook.xml 이 정한 시트 순서 그대로 시트 XML 경로.

    ``xl/worksheets/sheetN.xml`` 의 N 은 표시 순서와 무관하므로
    관계(rels)를 거쳐야 openpyxl 의 ``wb.worksheets`` 순서와 맞는다.
    """
    wb = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    target = {r.get("Id"): r.get("Target") for r in rels.findall(f"{_PKG_REL}Relationship")}
    out: list[str] = []
    for sheet in wb.findall(f"{_MAIN}sheets/{_MAIN}sheet"):
        path = target.get(sheet.get(f"{_REL}id"))
        if not path:
            continue
        path = path.lstrip("/")
        out.append(path if path.startswith("xl/") else f"xl/{path}")
    return out


def _vertical_merges(data: bytes) -> list[dict[int, list[int]]]:
    """시트 순서대로 {시작행: [열, ...]} — 세로 병합의 값 전파 대상.

    ``read_only=True`` 로 연 워크시트는 ``merged_cells`` 가 비어 있고, 끄면
    큰 시트에서 메모리가 위험하다(운영 최대 146,472셀). 병합 정보는 시트당
    수십~수백 개뿐이라 ZIP 에서 따로 읽는 편이 싸고, 실측 12건 전부
    openpyxl(read_only=False) 결과와 일치했다 — 그 중 1건은 openpyxl 이
    아예 열지 못한 파일이었다.

    가로 병합은 손대지 않는다. 지금도 좌상단에만 값이 있어 같은 줄 안 중복일
    뿐이고, HWP 쪽 격자 펼침 규칙과도 같다.
    """
    result: list[dict[int, list[int]]] = []
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            names = set(zf.namelist())
            for path in _sheet_xml_paths(zf):
                spans: dict[int, list[int]] = {}
                if path not in names:
                    result.append(spans)
                    continue
                with zf.open(path) as fh:
                    for _event, elem in ET.iterparse(fh, events=("end",)):
                        if elem.tag == f"{_MAIN}mergeCell":
                            ref = elem.get("ref") or ""
                            if ":" in ref:
                                a, b = ref.split(":", 1)
                                ma, mb = _CELL_REF.match(a), _CELL_REF.match(b)
                                if ma and mb:
                                    r1, r2 = int(ma.group(2)), int(mb.group(2))
                                    col = _col_num(ma.group(1))
                                    if r2 > r1:  # 세로로 걸친 것만
                                        for r in range(r1 + 1, r2 + 1):
                                            spans.setdefault(r, []).append(col)
                        elem.clear()  # 시트 본문을 메모리에 쌓지 않는다
                result.append(spans)
    except Exception as exc:  # noqa: BLE001
        logger.info("xlsx 병합 정보 읽기 실패(전파 생략): %s", exc)
        return []
    return result


def _sheet_rows(ws, budget: list[int], vmerge: dict[int, list[int]] | None = None) -> list[list[str]]:
    """시트를 문자열 행 목록으로. 빈 행·오른쪽 빈 열은 떨군다.

    세로 병합된 분류 값은 아래 행으로 복제한다 — 행이 스스로를 설명해야
    청크가 잘려도 소속을 잃지 않는다(HWP 쪽 rowspan 처리와 같은 규칙).
    """
    rows: list[list[str]] = []
    carry: dict[int, str] = {}  # 1-based 열 -> 위에서 내려온 값
    start = getattr(ws, "min_row", 1) or 1
    for row_idx, raw in enumerate(ws.iter_rows(values_only=True), start=start):
        if budget[0] <= 0:
            break
        cells = ["" if v is None else _fmt(v) for v in raw]
        budget[0] -= len(raw)

        for col in (vmerge or {}).get(row_idx, ()):
            value = carry.get(col)
            if not value:
                continue
            while len(cells) < col:
                cells.append("")
            if not cells[col - 1]:
                cells[col - 1] = value

        for i, value in enumerate(cells):
            if value:
                carry[i + 1] = value

        while cells and not cells[-1]:
            cells.pop()
        if not cells:
            continue
        rows.append(cells)
    return rows


def _to_table(rows: list[list[str]]) -> list[str]:
    """행 목록 → 마크다운 표. 첫 행을 헤더로 쓴다."""
    width = max(len(r) for r in rows)
    padded = [r + [""] * (width - len(r)) for r in rows]
    head, *body = padded
    lines = [
        "| " + " | ".join(head) + " |",
        "| " + " | ".join(["---"] * width) + " |",
    ]
    lines.extend("| " + " | ".join(r) + " |" for r in body)
    return lines


def xlsx_to_markdown(
    data: bytes, *, max_cells: int = MAX_CELLS, max_bytes: int = MAX_BYTES
) -> str:
    """XLSX 바이트를 마크다운으로. 열 수 없으면 XlsxParseError.

    시트가 둘 이상이면 시트명을 소제목으로 단다. 하나뿐이면 파일명이 이미
    제목으로 붙으므로 생략한다(`Sheet1` 같은 기본명이 본문에 섞이는 것 방지).

    셀 수(max_cells)와 출력 바이트(max_bytes) 양쪽으로 자른다. 둘 중 하나만
    막으면 나머지로 새어 한도를 넘긴다.
    """
    if data[:4] == _OLE2_MAGIC:
        raise XlsxParseError(
            "OLE2 컨테이너 — 암호가 걸렸거나 구형 이진 포맷이라 열 수 없음"
        )

    wb = _load(data)
    try:
        sheets = wb.worksheets
        vmerges = _vertical_merges(data)
        budget = [max_cells]
        chunks: list[str] = []
        used = 0
        truncated = False
        for idx, ws in enumerate(sheets):
            rows = _sheet_rows(ws, budget, vmerges[idx] if idx < len(vmerges) else None)
            if not rows:
                continue
            lines = ([f"## {ws.title}"] if len(sheets) > 1 else []) + _to_table(rows)
            for line in lines:
                size = len(line.encode("utf-8")) + 1
                if used + size > max_bytes:
                    truncated = True
                    break
                chunks.append(line)
                used += size
            chunks.append("")
            if truncated:
                break
        truncated = truncated or budget[0] <= 0
    finally:
        wb.close()

    if not chunks or not any(c.strip() for c in chunks):
        raise XlsxParseError("내용 없는 스프레드시트(빈 셀만)")

    if truncated:
        chunks.append("> (크기 상한에서 잘림 — 원본 일부만 색인됨)")
    return "\n".join(chunks).strip()
